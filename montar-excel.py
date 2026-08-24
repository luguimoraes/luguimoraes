#!/usr/bin/env python3
"""Monta o Excel de duas abas da reativacao de contatos.

Uso:
    python3 montar-excel.py export.csv selecao.csv [contatos-reativacao.xlsx]

    export.csv    export de contatos do SenseData, com as 41 colunas
    selecao.csv   saida de abas-reativacao.sql: chave + veredito

Saem:

    contatos-reativacao.xlsx   uma aba de conferencia e uma por lote
    manutencao-lote-A.csv      sobe agora
    manutencao-lote-B.csv      so depois do piloto

As abas tem papeis diferentes e nao sao o mesmo recorte de colunas:

    Inativos (de-para)   as 41 colunas do export + as de controle da selecao.
                         E a aba de conferencia: quem ficou inativo e por que.
                         O 'Ativo' aqui e o estado ATUAL, ou seja, False.

    Manutencao lote A    tres colunas: 'Cliente (ID Original)', 'E-mail' e
    Manutencao lote B    'Ativo' = Sim. Sao os obrigatorios da manutencao mais
                         o que muda; como ela grava so as colunas presentes,
                         tres nao tem como apagar Cargo, Telefone ou
                         Benchmarking de ninguem.

                         No lote A a dupla (Cliente, E-mail) acha um contato
                         so — sem ambiguidade, sobe direto. No lote B ela acha
                         mais de um, e o resultado depende de a manutencao
                         casar respeitando a caixa do e-mail ou nao. Quem
                         separa os dois e a coluna '_Lote' da selecao.

O arquivo que sobe sai da SELECAO, nao do export: se o export vier incompleto,
a aba de conferencia sai furada mas a reativacao continua inteira.

CUIDADO: a manutencao identifica por (Cliente, E-mail), nao pelo ID do
contato. Duas linhas com o mesmo e-mail na mesma conta — a forma do alvo —
ela nao separa. O e-mail sai com a caixa exata do registro escolhido; se ela
casar ignorando maiuscula/minuscula, ativa o par inteiro. Suba uma conta
primeiro.

Os dois arquivos precisam da coluna 'ID Contato' — nao para a manutencao, que
nao a usa, mas para o script casar export e selecao na aba de conferencia sem
confundir os dois registros do par.

O cabecalho do export sai byte a byte como entrou. 'Brand ' e 'Produto ' tem
espaco no fim de verdade; se o script aparar, a manutencao nao reconhece.

Tudo vai como TEXTO no xlsx: Excel nao transforma '132626-TAX' em outra coisa,
nem corta zero a esquerda, nem reinterpreta data.
"""

import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ABA_INATIVOS = "Inativos (de-para)"
ABA_MANUTENCAO = "Manutencao lote {}"

COL_ID = "ID Contato"
COL_CONTA = "ID Original"
COL_EMAIL = "Email"
COL_ENTRA = "_Entra no arquivo"
COL_LOTE = "_Lote"

# cabecalho do arquivo que sobe: os obrigatorios da manutencao + o que muda
CAB_MANUTENCAO = ["Cliente (ID Original)", "E-mail", "Ativo"]
VALOR_ATIVO = "Sim"
VERDADEIRO = {"t", "true", "sim", "s", "yes", "y", "1"}

LARGURA_MIN, LARGURA_MAX, LINHAS_PARA_MEDIR = 10, 46, 400


def ler_csv(caminho):
    """Le um CSV/TSV descobrindo o separador. Devolve (cabecalho, linhas)."""
    bruto = Path(caminho).read_text(encoding="utf-8-sig")
    if not bruto.strip():
        raise SystemExit(f"{caminho}: arquivo vazio")
    try:
        dialeto = csv.Sniffer().sniff(bruto[:8192], delimiters=",;\t|")
    except csv.Error:
        dialeto = csv.excel
    linhas = [l for l in csv.reader(bruto.splitlines(), dialeto) if any(c.strip() for c in l)]
    if not linhas:
        raise SystemExit(f"{caminho}: nenhuma linha util")
    return linhas[0], linhas[1:]


def indice(cabecalho, nome, onde):
    for i, coluna in enumerate(cabecalho):
        if coluna.strip() == nome:
            return i
    raise SystemExit(f"{onde}: falta a coluna {nome!r}")


def campo(linha, i):
    return linha[i].strip() if i < len(linha) else ""


def larguras(cabecalho, linhas):
    larg = [len(t) for t in cabecalho]
    for linha in linhas[:LINHAS_PARA_MEDIR]:
        for i, valor in enumerate(linha):
            if i < len(larg):
                larg[i] = max(larg[i], len(valor))
    return [min(max(n + 2, LARGURA_MIN), LARGURA_MAX) for n in larg]


def montar_aba(ws, titulo, cabecalho, linhas):
    ws.title = titulo
    ws.append(cabecalho)
    fundo = PatternFill("solid", fgColor="1F3864")
    for celula in ws[1]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = fundo

    for linha in linhas:
        ws.append([(linha[i] if i < len(linha) else "") for i in range(len(cabecalho))])
    for linha in ws.iter_rows(min_row=2):
        for celula in linha:
            celula.data_type = "s"
            celula.alignment = Alignment(vertical="top")
    for i, largura in enumerate(larguras(cabecalho, linhas), start=1):
        ws.column_dimensions[get_column_letter(i)].width = largura

    ws.freeze_panes = "A2"
    if linhas:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cabecalho))}{len(linhas) + 1}"


def main(argv):
    if len(argv) < 3:
        raise SystemExit(__doc__)

    saida = Path(argv[3] if len(argv) > 3 else "contatos-reativacao.xlsx")
    cab_export, lin_export = ler_csv(argv[1])
    cab_selecao, lin_selecao = ler_csv(argv[2])

    i_id_exp = indice(cab_export, COL_ID, "export")
    i_id_sel = indice(cab_selecao, COL_ID, "selecao")
    i_conta = indice(cab_selecao, COL_CONTA, "selecao")
    i_email = indice(cab_selecao, COL_EMAIL, "selecao")
    i_entra = indice(cab_selecao, COL_ENTRA, "selecao")
    i_lote = indice(cab_selecao, COL_LOTE, "selecao")
    i_controle = [i for i, c in enumerate(cab_selecao) if c.strip().startswith("_")]
    controle = [cab_selecao[i].strip() for i in i_controle]

    veredito = {campo(l, i_id_sel): l for l in lin_selecao}
    if len(veredito) != len(lin_selecao):
        raise SystemExit(f"selecao: {COL_ID} repetido — a selecao devia ter um por contato")

    # O que sobe sai da selecao, nao do export: um export incompleto fura a aba
    # de conferencia, mas nao pode encolher a reativacao.
    lotes = {}
    for linha in lin_selecao:
        if campo(linha, i_entra).lower() not in VERDADEIRO:
            continue
        alvo = lotes.setdefault(campo(linha, i_lote) or "?", [])
        alvo.append([campo(linha, i_conta), campo(linha, i_email), VALOR_ATIVO])
    for linhas in lotes.values():
        linhas.sort(key=lambda l: (l[0], l[1].lower()))

    # A manutencao casa por (Cliente, E-mail) e nao enxerga caixa nem ID. Duas
    # linhas com a mesma dupla mandariam ordens ambiguas para o mesmo par.
    subindo = [l for linhas in lotes.values() for l in linhas]
    pares = {(l[0], l[1].lower()) for l in subindo}
    if len(pares) != len(subindo):
        raise SystemExit(
            f"selecao: {len(subindo) - len(pares)} linhas repetem "
            f"(Cliente, E-mail) no arquivo que sobe.\n"
            "A manutencao nao separa duas linhas da mesma dupla — a selecao\n"
            "tem que escolher UM registro por pessoa-conta antes de chegar aqui."
        )

    cab_inativos = cab_export + controle
    lin_inativos = []
    for linha in lin_export:
        sel = veredito.get(campo(linha, i_id_exp))
        if sel is not None:
            lin_inativos.append(linha + [(sel[i] if i < len(sel) else "") for i in i_controle])

    wb = Workbook()
    montar_aba(wb.active, ABA_INATIVOS, cab_inativos, lin_inativos)
    for lote in sorted(lotes):
        montar_aba(wb.create_sheet(), ABA_MANUTENCAO.format(lote),
                   CAB_MANUTENCAO, lotes[lote])
    wb.save(saida)

    print(f"{saida}")
    print(f"  {ABA_INATIVOS:<22} {len(lin_inativos):>6} linhas · {len(cab_inativos)} colunas")
    for lote in sorted(lotes):
        aba = ABA_MANUTENCAO.format(lote)
        print(f"  {aba:<22} {len(lotes[lote]):>6} linhas · {len(CAB_MANUTENCAO)} colunas")

    print()
    for lote in sorted(lotes):
        arquivo = saida.with_name(f"manutencao-lote-{lote}.csv")
        with open(arquivo, "w", encoding="utf-8-sig", newline="") as f:
            escritor = csv.writer(f)
            escritor.writerow(CAB_MANUTENCAO)
            escritor.writerows(lotes[lote])
        quando = "sobe agora" if lote == "A" else "so depois do piloto"
        print(f"{arquivo}  <- {quando}")

    faltando = len(veredito) - len(lin_inativos)
    if faltando:
        print()
        print(f"  ATENCAO: {faltando} contatos da selecao nao estao no export.")
        print(f"  So a aba '{ABA_INATIVOS}' sai incompleta — o arquivo que sobe")
        print("  vem da selecao e continua inteiro. Reexporte incluindo os inativos.")


if __name__ == "__main__":
    main(sys.argv)
